from cgitb import reset
import os
import sys
from openpyxl import Workbook, load_workbook
#Otiweramy nasz plik i tworzymy uchwyt do niego
workbook=load_workbook(filename="mail_database.xlsx")
sheet = workbook.active
sheet.title="Data"

class User():
    #parametry użyte przy logowaniu
    def __init__(self, mail="",password="", name="",year = 0, phone = 0,country="",city = "" ):
        self.name=name
        self.mail=mail
        self.password=password
        self.year=year
        self.phone = phone
        self.country=country
        self.city = city
    #rejestrowanie
    def register(self):
        print("ZAREJESTRUJ SIĘ: ")
        #podajemy nazwe uzytkownika
        self.mail = input("Podaj swój adres e-mail: ")
        #sprawdzamy czy nazwa znajduje sie juz w bazie w tym celu przeszukujemy kolumne A
        for row in sheet.iter_rows():
            for cell in sheet['A']:
                #Jeżeli nasz mail znajduje juz sie w bazie to odsylamy do logowania
                if cell.value == self.mail:

                    print("Dany użytkownik już istnieje. Proszę się zalogować.")
                    self.loggin()
                    
                
                
                 
        # prosimy o podanie pozostalych informacji
        self.password = input("Podaj hasło: ")
        pass_again = input("Powtórz hasło: ")
        if self.password != pass_again:
            print("Wpowadzone hasła różnią się od siebie. Spróbuj ponownie.")
            self.register()

        else:
            self.name = input("Podaj swoje imię: ")
            self.year = input("Podaj swój rok urodzenia: ")
            self.phone = input("Podaj numer telefonu: ")
            self.country = input("Z jakiego kraju jesteś: ")
            self.city = input("Z jakiego miasta jesteś: ")
            #dodajemy dane do exela w danej kolejności i zapisujemy
            sheet.append([self.mail, self.password,self.name, self.year, self.phone, self.country, self.city])
            workbook.save(filename="mail_database.xlsx")
            print(self.name+", witaj na naszym serwisie. Przed tobą wiele przygód!")
    #logowanie
    def loggin(self):
        print("LOGOWANIE")
        log = False
        self.mail = input("Podaj swój adres e-mail: ")
        #szukamy komorki w kolumnie A gdzie znajduje sie nasz email
        for row in sheet.iter_rows():
            for cell in sheet['A']:
                if cell.value == self.mail:
                    #jezeli znalezlismy ten sam adres sprawdzamy haslo w tym celu pobieramy informacje w ktorym rzedzie znajduje sie mail
                   log=True
                   c_row=int(cell.row)
                   m_password=input("Podaj hasło: ")
                   #haslo znajduje sie w tym samym rzedzie ale w kolumnie obok dlatego dajemy 2
                   if(m_password==sheet.cell(c_row,2).value):

                       print("Hasło poprawne")
                       print("Witaj "+sheet.cell(c_row,3).value)
                       main()
                   else:
                       print("Haslo niepoprawne!")
                       self.loggin()



        if log == False:        
            print("Brak użytkownika w bazie. Proszę się zarejestrować.")
            self.register()
        



          
def main():
    while True:
        opt=input("1.Zaloguj się/2.Zarejestruj użytkownika/3.Wyjdź: ")
        u = User()
        if int(opt)==1:
            u.loggin()
        elif int(opt)==2:
            u.register()
        elif int(opt)==3:
            quit()
        else:
            continue
    
main()


